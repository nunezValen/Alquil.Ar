from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login, update_session_auth_hash, logout
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.utils import timezone
from django.conf import settings
from django.views.decorators.csrf import csrf_protect, ensure_csrf_cookie, csrf_exempt
from django.views.decorators.http import require_http_methods, require_GET
from django.urls import reverse
from .models import Persona, Maquina, Sucursal, CodigoVerificacion
from .forms import (
    PersonaForm, ClienteForm, EmpleadoForm, EditarPersonaForm,
    CambiarPasswordForm, ModificarDatosPersonalesForm
)
from datetime import date
import random
import string
import mercadopago
import binance
from pyngrok import ngrok
import json
from .utils import generar_password_random
from django.db.models import Q, Count, Avg, F, ExpressionWrapper, fields
from django.core.paginator import Paginator
from maquinas.models import Alquiler, MaquinaBase
from maquinas.utils import enviar_email_alquiler_simple, enviar_email_alquiler_cancelado
from django.db.models.functions import Coalesce
from django.db.models import Value
from functools import wraps
from django.utils.dateparse import parse_date

def empleado_requerido(view_func):
    """
    Decorador que verifica que el usuario sea un empleado activo o un superusuario.
    Si no cumple, muestra un mensaje de error y redirige a la página de inicio.
    """
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.error(request, "Debes iniciar sesión para ver esta página.")
            return redirect('persona:login_unificado2')

        if request.user.is_superuser:
            return view_func(request, *args, **kwargs)

        try:
            persona = Persona.objects.get(email=request.user.email)
            if persona.es_empleado:
                if persona.bloqueado_empleado:
                    messages.error(request, 'Tu rol de empleado está bloqueado. No puedes acceder a esta sección.')
                    return redirect('persona:inicio')
                else:
                    return view_func(request, *args, **kwargs)
            else:
                messages.error(request, 'No tienes los permisos de empleado necesarios para acceder a esta página.')
                return redirect('persona:inicio')
        except Persona.DoesNotExist:
            messages.error(request, 'No se encontró un perfil de usuario asociado. Contacta al administrador.')
            return redirect('persona:inicio')
    return _wrapped_view

def es_admin(user):
    """
    Verifica si el usuario es administrador
    """
    return user.is_authenticated and user.is_superuser

def es_empleado_o_admin(user):
    """Verifica si el usuario es empleado (y no está bloqueado) o administrador"""
    if not user.is_authenticated:
        return False
        
    # El superusuario (admin) siempre tiene acceso
    if user.is_superuser:
        return True
        
    try:
        persona = Persona.objects.get(email=user.email)
        # Un empleado tiene acceso si tiene el rol Y NO está bloqueado
        return persona.es_empleado and not persona.bloqueado_empleado
    except Persona.DoesNotExist:
        # Si no tiene perfil de Persona, no puede ser empleado
        return False

def inicio(request):
    """Vista de inicio con manejo de retorno de MercadoPago"""
    
    # NOTA: La lógica para 'empleados_emails' y 'es_empleado_actuando_como_cliente'
    # ahora es manejada por el context_processor para asegurar consistencia.
    # Ya no se definen localmente en esta vista.
    
    # Manejar retorno de MercadoPago (fallback si el webhook no funciona)
    status = request.GET.get('status')
    external_reference = request.GET.get('external_reference')
    
    if status and external_reference and request.user.is_authenticated:
        print(f"=== PROCESANDO RETORNO DE MERCADOPAGO ===")
        print(f"Status: {status}")
        print(f"External reference: {external_reference}")
        
        if status == 'approved':
            try:
                # Parsear los datos del external_reference
                # Formato: "maquina_id|persona_id|fecha_inicio|fecha_fin|metodo_pago|total"
                datos = external_reference.split('|')
                if len(datos) == 6:
                    from datetime import datetime
                    
                    maquina_id, persona_id, fecha_inicio_str, fecha_fin_str, metodo_pago, total = datos
                    
                    # Convertir fechas
                    fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
                    fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
                    
                    # Obtener objetos
                    maquina_base = MaquinaBase.objects.get(id=maquina_id)
                    persona = Persona.objects.get(id=persona_id)
                    
                    # Verificar que el usuario actual es el mismo que hizo el pago
                    persona_actual = Persona.objects.get(email=request.user.email)
                    if persona.id == persona_actual.id:
                        
                        # Verificar que no existe ya un alquiler para estos datos
                        alquiler_existente = Alquiler.objects.filter(
                            persona=persona,
                            maquina_base=maquina_base,
                            fecha_inicio=fecha_inicio,
                            fecha_fin=fecha_fin
                        ).first()
                        
                        if not alquiler_existente:
                            # Verificar disponibilidad nuevamente
                            if Alquiler.verificar_disponibilidad(maquina_base, fecha_inicio, fecha_fin):
                                # Verificar que el cliente no tenga otro alquiler activo
                                alquileres_activos = Alquiler.objects.filter(
                                    persona=persona,
                                    estado__in=['reservado', 'en_curso']
                                )
                                
                                if not alquileres_activos.exists():
                                    # Crear el alquiler
                                    alquiler = Alquiler.objects.create(
                                        maquina_base=maquina_base,
                                        persona=persona,
                                        fecha_inicio=fecha_inicio,
                                        fecha_fin=fecha_fin,
                                        metodo_pago=metodo_pago,
                                        estado='reservado',
                                        monto_total=float(total),
                                        preference_id=request.GET.get('preference_id')
                                    )
                                    
                                    print(f"=== ALQUILER CREADO VIA RETORNO ===")
                                    print(f"ID: {alquiler.id}")
                                    print(f"Número: {alquiler.numero}")
                                    print(f"Código de retiro: {alquiler.codigo_retiro}")
                                    
                                    # Enviar email con PDF al cliente
                                    try:
                                        print(f"[INFO] Intentando enviar email desde persona webhook...")
                                        resultado_email = enviar_email_alquiler_simple(alquiler)
                                        if resultado_email:
                                            print(f"[SUCCESS] Email enviado correctamente desde persona webhook")
                                        else:
                                            print(f"[ERROR] Falló el envío de email desde persona webhook")
                                    except Exception as e:
                                        print(f"[ERROR] Error al enviar email desde persona webhook: {str(e)}")
                                        import traceback
                                        traceback.print_exc()
                                    
                                    messages.success(request, f'¡Pago exitoso! Tu número de alquiler es: {alquiler.numero}. Código de retiro: {alquiler.codigo_retiro}')
                                else:
                                    print(f"Alquiler ya existe: {alquiler_existente.numero}")
                                    messages.success(request, f'Alquiler ya confirmado: {alquiler_existente.numero}')
            except Exception as e:
                print(f"Error al procesar retorno de pago: {str(e)}")
                messages.error(request, 'Hubo un error al procesar el pago.')

    # Obtener máquinas para mostrar en la página de inicio
    maquinas = MaquinaBase.objects.filter(stock__gt=0)[:4]  # Obtener las primeras 4 máquinas con stock
    for maquina in maquinas:
        if len(maquina.descripcion_corta) > 100:
            maquina.descripcion_vista = maquina.descripcion_corta[:100] + "..."
        else:
            maquina.descripcion_vista = maquina.descripcion_corta
     
    return render(request, 'persona/inicio.html', {
        'maquinas': maquinas,
    })

@login_required
def catalogo_maquinas(request):
    # Usar el sistema nuevo de MaquinaBase con unidades
    from maquinas.models import MaquinaBase
    # Filtrar solo máquinas que tengan unidades realmente disponibles (no en mantenimiento)
    maquinas = MaquinaBase.objects.filter(
        visible=True,
        unidades__estado='disponible',
        unidades__visible=True
    ).distinct().order_by('nombre')
    return render(request, 'persona/catalogo_maquinas.html', {'maquinas': maquinas})

@login_required
def detalle_maquina(request, maquina_id):
    maquina = get_object_or_404(Maquina, id=maquina_id)
    alquiler = None
    
    if request.method == 'POST':
        print(">>> POST recibido en detalle_maquina")
        form = AlquilerForm(request.POST, maquina=maquina)
        if form.is_valid():
            print(">>> Formulario válido")
            fecha_inicio = form.cleaned_data['fecha_inicio']
            fecha_fin = form.cleaned_data['fecha_fin']
            metodo_pago = form.cleaned_data['metodo_pago']
            
            # Validar días mínimos
            dias = (fecha_fin - fecha_inicio).days + 1
            if dias < maquina.dias_minimos:
                form.add_error(None, f"El alquiler mínimo es de {maquina.dias_minimos} días.")
            else:
                try:
                    # Obtener la persona asociada al usuario
                    try:
                        persona = Persona.objects.get(email=request.user.email)
                        print(f">>> Persona encontrada: {persona.email}")
                    except Persona.DoesNotExist:
                        print(">>> Error: No se encontró la persona")
                        form.add_error(None, "No se encontró tu perfil de persona. Por favor, regístrate primero.")
                        return render(request, 'persona/detalle_maquina.html', {
                            'maquina': maquina,
                            'form': form,
                            'mercadopago_public_key': settings.MERCADOPAGO_PUBLIC_KEY
                        })

                    # Crear el alquiler
                    alquiler = Alquiler.objects.create(
                        maquina=maquina,
                        persona=persona,
                        fecha_inicio=fecha_inicio,
                        fecha_fin=fecha_fin,
                        metodo_pago=metodo_pago,
                        estado='pendiente'
                    )
                    print(f">>> Alquiler creado con ID: {alquiler.id}")
                    
                    # Procesar el pago según el método seleccionado
                    if metodo_pago == 'mercadopago':
                        print(">>> Iniciando proceso de pago con Mercado Pago")
                        # Inicializar SDK de Mercado Pago
                        sdk = mercadopago.SDK(settings.MERCADOPAGO_ACCESS_TOKEN)
                        print(f">>> SDK inicializado con token: {settings.MERCADOPAGO_ACCESS_TOKEN[:10]}...")
                        
                        # Configurar preferencia de pago
                        base_url = request.build_absolute_uri('/').rstrip('/')
                        preference_data = {
                            "items": [{
                                "title": f"Alquiler de {maquina.nombre}",
                                "quantity": 1,
                                "currency_id": "ARS",
                                "unit_price": float(maquina.precio_dia * dias + 1)
                            }],
                            "back_urls": {
                                "success": f"{base_url}/persona/checkout/{alquiler.id}/?status=approved",
                                "failure": f"{base_url}/persona/checkout/{alquiler.id}/?status=failure",
                                "pending": f"{base_url}/persona/checkout/{alquiler.id}/?status=pending"
                            },
                            "auto_return": "approved",
                            "external_reference": str(alquiler.id),
                            "notification_url": f"{base_url}/persona/webhook/mercadopago/",
                            "binary_mode": True,
                            "payer": {
                                "email": request.user.email
                            },
                            "payment_methods": {
                                "excluded_payment_types": [
                                    {"id": "ticket"}
                                ],
                                "installments": 1
                            }
                        }
                        print(f">>> Datos de preferencia: {preference_data}")
                        
                        # Crear preferencia
                        preference_response = sdk.preference().create(preference_data)
                        print(f">>> Respuesta de Mercado Pago: {preference_response}")
                        
                        if preference_response["status"] == 201:
                            preference = preference_response["response"]
                            print(f">>> Preferencia creada con ID: {preference['id']}")
                            print(f">>> URL de redirección: {preference['init_point']}")
                            
                            # Guardar el ID de preferencia
                            alquiler.preference_id = preference["id"]
                            alquiler.save()
                            
                            # Retornar la URL de MercadoPago
                            return JsonResponse({
                                'status': 'success',
                                'redirect_url': preference["init_point"]
                            })
                        else:
                            error_msg = f"Error al crear preferencia: {preference_response.get('message', 'Error desconocido')}"
                            print(f">>> {error_msg}")
                            raise Exception(error_msg)
                        
                    elif metodo_pago == 'binance':
                        # Implementar lógica de Binance Pay
                        pass
                        
                except Exception as e:
                    print(f">>> Error en el proceso: {str(e)}")
                    if alquiler:
                        alquiler.delete()
                    form.add_error(None, f"Error al procesar el pago: {str(e)}")
                    return JsonResponse({
                        'status': 'error',
                        'message': str(e)
                    }, status=400)
        else:
            print(f">>> Errores en el formulario: {form.errors}")
            return JsonResponse({
                'status': 'error',
                'message': 'Formulario inválido',
                'errors': dict(form.errors)
            }, status=400)
    else:
        form = AlquilerForm(maquina=maquina)
    
    return render(request, 'persona/detalle_maquina.html', {
        'maquina': maquina,
        'form': form,
        'mercadopago_public_key': settings.MERCADOPAGO_PUBLIC_KEY
    })

@login_required
def mis_alquileres(request):
    """
    Vista para que los usuarios vean su historial de alquileres
    Accesible para clientes y empleados
    """
    # Los empleados y clientes pueden ver sus alquileres
    es_empleado_actuando_como_cliente = request.session.get('es_empleado_actuando_como_cliente', False)
    
    try:
        persona = Persona.objects.get(email=request.user.email)
        
        # Verificar que el usuario tenga el rol de cliente activo
        if not persona.es_cliente or persona.bloqueado_cliente:
            messages.error(request, 'No tienes acceso a esta sección porque tu rol de cliente no está activo.')
            return redirect('persona:inicio')
        
        # Obtener todos los alquileres del cliente, ordenados por fecha de creación
        alquileres = Alquiler.objects.filter(persona=persona).order_by('-fecha_creacion')
        
        # Agregar información adicional para cada alquiler
        for alquiler in alquileres:
            # Calcular días restantes si está en curso o reservado
            if alquiler.estado in ['reservado', 'en_curso'] and alquiler.fecha_inicio:
                from datetime import date
                dias_hasta_inicio = (alquiler.fecha_inicio - date.today()).days
                if dias_hasta_inicio > 0:
                    alquiler.dias_hasta_inicio = dias_hasta_inicio
                elif alquiler.fecha_fin:
                    dias_hasta_fin = (alquiler.fecha_fin - date.today()).days
                    if dias_hasta_fin >= 0:
                        alquiler.dias_hasta_fin = dias_hasta_fin
            
            # Calcular reembolso potencial si puede cancelar
            if alquiler.puede_ser_cancelado():
                porcentaje, monto = alquiler.calcular_reembolso(es_empleado=False)
                alquiler.porcentaje_reembolso_cliente = porcentaje
                alquiler.monto_reembolso_cliente = monto
        
        # Calcular estadísticas del resumen
        from django.db.models import Count, Q
        stats = {
            'total': alquileres.count(),
            'activos': alquileres.filter(estado__in=['reservado', 'en_curso']).count(),
            'finalizados': alquileres.filter(estado='finalizado').count(),
            'cancelados': alquileres.filter(estado='cancelado').count()
        }
        
        return render(request, 'persona/mis_alquileres.html', {
            'alquileres': alquileres,
            'persona': persona,
            'stats': stats
        })
        
    except Persona.DoesNotExist:
        messages.error(request, 'No se encontró tu perfil de cliente.')
        return redirect('persona:inicio')

@login_required 
def cancelar_mi_alquiler(request, alquiler_id):
    """
    Vista para que los clientes cancelen sus propios alquileres
    Accesible para clientes y empleados actuando como clientes
    """
    # Verificar que el usuario sea cliente o empleado actuando como cliente
    es_empleado_actuando_como_cliente = request.session.get('es_empleado_actuando_como_cliente', False)
    if es_empleado_o_admin(request.user) and not es_empleado_actuando_como_cliente:
        messages.error(request, 'Los empleados no pueden usar esta función. Usa la gestión de alquileres.')
        return redirect('persona:gestion')
    
    try:
        persona = Persona.objects.get(email=request.user.email)
        
        # Verificar que el cliente tenga el rol activo
        if not persona.es_cliente or persona.bloqueado_cliente:
            messages.error(request, 'No puedes cancelar alquileres porque tu rol de cliente no está activo.')
            return redirect('persona:inicio')

        alquiler = get_object_or_404(Alquiler, id=alquiler_id, persona=persona)
        
        if request.method == 'POST':
            try:
                if not alquiler.puede_ser_cancelado():
                    messages.error(request, f'El alquiler {alquiler.numero} no puede ser cancelado (estado: {alquiler.get_estado_display()})')
                    return redirect('persona:mis_alquileres')
                
                observaciones = request.POST.get('observaciones', 'Cancelado por el cliente')
                porcentaje, monto = alquiler.cancelar(empleado=None, observaciones=observaciones)
                
                # Enviar email de cancelación
                try:
                    enviar_email_alquiler_cancelado(alquiler)
                    print(f"[INFO] Email de cancelación enviado para alquiler {alquiler.numero}")
                except Exception as e:
                    print(f"[ERROR] Error al enviar email de cancelación: {str(e)}")
                
                if porcentaje > 0:
                    messages.success(request, 
                        f'Alquiler {alquiler.numero} cancelado exitosamente. '
                        f'Tienes derecho a un reembolso del {porcentaje}% (${monto:.2f}). '
                        f'Acércate a la tienda el próximo mes para cobrarlo. '
                        f'Recibirás un email de confirmación.')
                else:
                    messages.success(request, 
                        f'Alquiler {alquiler.numero} cancelado exitosamente. '
                        f'No corresponde reembolso según la política de cancelación. '
                        f'Recibirás un email de confirmación.')
                
            except Exception as e:
                messages.error(request, f'Error al cancelar el alquiler: {str(e)}')
            
            return redirect('persona:mis_alquileres')
        
        # Calcular reembolso para mostrar en el template
        porcentaje, monto = alquiler.calcular_reembolso(es_empleado=False)
        
        return render(request, 'persona/cancelar_mi_alquiler.html', {
            'alquiler': alquiler,
            'porcentaje_reembolso': porcentaje,
            'monto_reembolso': monto,
            'persona': persona
        })
        
    except Persona.DoesNotExist:
        messages.error(request, 'No se encontró tu perfil de cliente.')
        return redirect('persona:inicio')

def lista_maquinas(request):
    search_query = request.GET.get('q', '')
    maquinas = MaquinaBase.objects.filter(stock__gt=0)  # Solo máquinas con stock

    if search_query:
        maquinas = maquinas.filter(nombre__icontains=search_query)

    for maquina in maquinas:
        if len(maquina.descripcion_corta) > 200:
            maquina.descripcion_vista = maquina.descripcion_corta[:197] + "..."
        else:
            maquina.descripcion_vista = maquina.descripcion_corta

    return render(request, 'persona/catalogo.html', {'maquinas': maquinas})

def lista_empleados(request):
    empleados = Persona.objects.filter(es_empleado=True)  # Consulta todos los empleados
    return render(request, 'lista_empleado.html', {'empleados': empleados})

def lista_personas(request):
    personas = Persona.objects.all()
    personas = Persona.objects.all()  # Consulta todas las personas
    return render(request, 'lista_persona.html', {'personas': personas})

@empleado_requerido
@csrf_protect
@ensure_csrf_cookie
@require_http_methods(["GET", "POST"])
def registrar_persona(request):
    # Redirigir a la página de login ya que el registro general no está permitido
    messages.info(request, 'El registro de usuarios solo puede ser realizado por empleados o administradores.')
    return redirect('persona:login_unificado2')

@csrf_protect
@ensure_csrf_cookie
@require_http_methods(["GET", "POST"])
def login_view(request):
    error = None
    if request.method == 'POST':
        email = request.POST.get('username')
        password = request.POST.get('password')

        try:
            persona = Persona.objects.get(email=email)
            user = authenticate(request, username=email, password=password)
            if user is not None:
                # La comprobación de "ambos bloqueados" ya existe y es correcta.
                if (((not persona.es_cliente) or persona.bloqueado_cliente) and ((not persona.es_empleado) or persona.bloqueado_empleado)):
                    error = 'Tu cuenta está suspendida. Contacta al administrador.'
                else:
                    login(request, user)
                    # Redirigir según el rol activo
                    if persona.es_empleado and not persona.bloqueado_empleado:
                        return redirect('persona:inicio')
                    elif persona.es_cliente and not persona.bloqueado_cliente:
                        return redirect('persona:inicio')
                    elif persona.es_admin:
                         return redirect('persona:inicio')
                    else:
                        error = 'Tu cuenta no está registrada como cuenta de empleado ni de cliente.'
            else:
                error = 'Email o contraseña incorrectos'
        except Persona.DoesNotExist:
            error = 'Email o contraseña incorrectos'

    return render(request, 'login.html', {'error': error})

@login_required
def pagina_principal(request):
    return redirect('persona:inicio')

@login_required
@user_passes_test(es_admin)
@csrf_protect
@ensure_csrf_cookie
@require_http_methods(["GET", "POST"])
def registrar_empleado(request):
    if request.method == 'POST':
        form = PersonaForm(request.POST)
        if form.is_valid():
            persona = form.save(commit=False)
            persona.es_empleado = True
            # El administrador decide si el empleado también es cliente
            persona.es_cliente = form.cleaned_data.get('es_cliente', False)
            persona.save()

            email = persona.email
            if email:
                if User.objects.filter(username=email).exists():
                    messages.error(request, 'Ya existe un usuario con ese email.')
                else:
                    try:
                        password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
                        user = User.objects.create_user(
                            username=email,
                            email=email,
                            password=password,
                            first_name=persona.nombre,
                            last_name=persona.apellido,
                            is_staff=True
                        )
                        send_mail(
                            'Tu cuenta de empleado en Alquil.ar',
                            f'Hola {persona.nombre},\n\nTu usuario ha sido creado.\n\nUsuario: {email}\nContraseña: {password}\n\nPor favor, cambia tu contraseña después de iniciar sesión.',
                            'no-reply@alquilar.com.ar',
                            [email],
                            fail_silently=False,
                        )
                        messages.success(request, 'Empleado registrado exitosamente. Recibirá su contraseña por email.')
                        return redirect('persona:registrar_empleado')
                    except Exception as e:
                        messages.error(request, f'Error al enviar el correo: {e}')
    else:
        form = PersonaForm()

    return render(request, 'persona/registrar_empleado.html', {'form': form})

@csrf_protect
@ensure_csrf_cookie
@require_http_methods(["GET", "POST"])
def login_empleado(request):
    error = None
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        # Agregamos el prefijo emp_ al username para buscar el usuario correcto
        user = authenticate(request, username=f"emp_{username}", password=password)
        if user is not None:
            login(request, user)
            return redirect('persona:inicio')
        else:
            error = 'Usuario o contraseña incorrectos.'
    return render(request, 'login_empleado.html', {'error': error})

@login_required
@csrf_protect
@ensure_csrf_cookie
@require_http_methods(["GET", "POST"])
def cambiar_password(request):
    if request.method == 'POST':
        paso = request.POST.get('paso')
        
        if paso == 'verificar':
            password_actual = request.POST.get('password_actual')
            
            # Verificar la contraseña actual
            user = request.user
            if not user.check_password(password_actual):
                messages.error(request, 'La contraseña actual es incorrecta.')
                return render(request, 'persona/cambiar_password.html', {'password_verificada': False})
            
            # Si la contraseña es correcta, mostrar el formulario de cambio
            return render(request, 'persona/cambiar_password.html', {'password_verificada': True})
            
        elif paso == 'cambiar':
            password_nuevo = request.POST.get('password_nuevo')
            password_confirmacion = request.POST.get('password_confirmar')
            
            # Validar longitud de la nueva contraseña
            if len(password_nuevo) < 6:
                messages.error(request, 'La nueva contraseña debe tener al menos 6 caracteres.')
                return render(request, 'persona/cambiar_password.html', {'password_verificada': True})
            
            if len(password_nuevo) > 16:
                messages.error(request, 'La nueva contraseña no puede tener más de 16 caracteres.')
                return render(request, 'persona/cambiar_password.html', {'password_verificada': True})
            
            # Validar que las contraseñas coincidan
            if password_nuevo != password_confirmacion:
                messages.error(request, 'Las contraseñas no coinciden.')
                return render(request, 'persona/cambiar_password.html', {'password_verificada': True})
            
            # Cambiar la contraseña
            user = request.user
            user.set_password(password_nuevo)
            user.save()
            
            # Actualizar la sesión para que el usuario no tenga que volver a iniciar sesión
            update_session_auth_hash(request, user)
            
            messages.success(request, 'Tu contraseña ha sido cambiada exitosamente.')
            return redirect('persona:inicio')
    
    return render(request, 'persona/cambiar_password.html', {'password_verificada': False})

@login_required
@csrf_protect
@ensure_csrf_cookie
@require_http_methods(["GET", "POST"])
def cambiar_password_empleado_logueado(request):
    error = None
    success = None
    password_verificada = False
    
    if request.method == 'POST':
        paso = request.POST.get('paso')
        
        if paso == 'verificar':
            password_actual = request.POST.get('password_actual', '')
            
            if not password_actual:
                error = 'Debe ingresar su contraseña actual.'
            else:
                # Verificar la contraseña actual
                user = request.user
                if not user.check_password(password_actual):
                    error = 'La contraseña actual es incorrecta.'
                else:
                    password_verificada = True
                    success = 'Contraseña verificada correctamente.'
        
        elif paso == 'cambiar':
            password_nuevo = request.POST.get('password_nuevo', '')
            password_confirmacion = request.POST.get('password_confirmacion', '')
            password_verificada = True  # Mantener el formulario de cambio visible

            if not password_nuevo or not password_confirmacion:
                error = 'Debe completar todos los campos.'
            elif password_nuevo != password_confirmacion:
                error = 'Las contraseñas no coinciden.'
            elif len(password_nuevo) < 6:
                error = 'La contraseña debe tener al menos 6 caracteres.'
            elif len(password_nuevo) > 16:
                error = 'La contraseña no puede tener más de 16 caracteres.'
            else:
                user = request.user
                user.set_password(password_nuevo)
                user.save()
                update_session_auth_hash(request, user)  # Mantiene la sesión activa
                messages.success(request, 'Contraseña cambiada exitosamente.')
                return redirect('persona:inicio')
    
    return render(request, 'persona/cambiar_password_empleado_logueado.html', {
        'error': error,
        'success': success,
        'password_verificada': password_verificada
    })

@login_required
@csrf_protect
@ensure_csrf_cookie
@require_http_methods(["GET", "POST"])
def cambiar_password_logueado(request):
    error = None
    success = None
    password_verificada = False
    
    if request.method == 'POST':
        paso = request.POST.get('paso')
        
        if paso == 'verificar':
            password_actual = request.POST.get('password_actual', '')
            
            if not password_actual:
                error = 'Debe ingresar su contraseña actual.'
            else:
                # Verificar la contraseña actual
                user = request.user
                if not user.check_password(password_actual):
                    error = 'La contraseña actual es incorrecta.'
                else:
                    password_verificada = True
        
        elif paso == 'cambiar':
            password1 = request.POST.get('password1', '')
            password2 = request.POST.get('password2', '')

            if not password1 or not password2:
                error = 'Debe completar todos los campos.'
                password_verificada = True
            elif password1 != password2:
                error = 'Las contraseñas no coinciden.'
                password_verificada = True
            elif len(password1) < 6:
                error = 'La contraseña debe tener al menos 6 caracteres.'
                password_verificada = True
            elif len(password1) > 16:
                error = 'La contraseña no puede tener más de 16 caracteres.'
                password_verificada = True
            else:
                user = request.user
                user.set_password(password1)
                user.save()
                update_session_auth_hash(request, user)  # Mantiene la sesión activa
                messages.success(request, 'Contraseña cambiada exitosamente.')
                return redirect('persona:inicio')
    
    return render(request, 'cambiar_password_logueado.html', {
        'error': error,
        'success': success,
        'password_verificada': password_verificada
    })

@login_required
@csrf_protect
@ensure_csrf_cookie
@require_http_methods(["GET", "POST"])
def login_as_persona(request):
    error = None
    success = None

    # Verificar que el usuario actual es un empleado
    try:
        persona = Persona.objects.get(email=request.user.email)
        # Solo empleados activos pueden usar esta función
        if not persona.es_empleado or persona.bloqueado_empleado:
            messages.error(request, 'Solo los empleados activos pueden acceder a esta función.')
            return redirect('persona:inicio')

        if request.method == 'POST':
            # Verificar que el empleado también tiene rol de cliente
            if not persona.es_cliente:
                messages.error(request, 'No tienes cuenta de cliente.')
                return redirect('persona:inicio')
            
            # Verificar que no esté bloqueado como cliente
            if persona.bloqueado_cliente:
                messages.error(request, 'No tienes permiso para actuar como cliente o tu rol de cliente está bloqueado.')
                return redirect('persona:inicio')

            # Verificar contraseña
            password = request.POST.get('password')
            if not password:
                error = 'Debes ingresar tu contraseña.'
            else:
                user = authenticate(request, username=request.user.email, password=password)
                if user is not None:
                    # Guardar en la sesión que el usuario es un empleado/admin actuando como cliente
                    request.session['es_empleado_actuando_como_cliente'] = True
                    messages.success(request, 'Has cambiado a tu cuenta personal exitosamente.')
                    return redirect('persona:inicio')
                else:
                    error = 'Contraseña incorrecta.'

    except Persona.DoesNotExist:
        messages.error(request, 'No se encontró tu perfil.')
        return redirect('persona:inicio')

    return render(request, 'persona/login_as_persona.html', {
        'error': error, 
        'success': success,
        'persona': persona
    })

@login_required
def inicio_blanco(request):
    maquinas = MaquinaBase.objects.filter(stock__gt=0)[:4]  # Obtener las primeras 4 máquinas con stock
    for maquina in maquinas:
        # Creamos un atributo temporal solo para la vista
        if len(maquina.descripcion_corta) > 300:
            maquina.descripcion_vista = maquina.descripcion_corta[:297] + "..."
        else:
            maquina.descripcion_vista = maquina.descripcion_corta
    return render(request, 'persona/inicio_blanco.html', {'maquinas': maquinas})

@csrf_protect
@ensure_csrf_cookie
@require_http_methods(["GET", "POST"])
def login_unificado2(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')

        # 1. Buscar en modelo Persona
        try:
            persona = Persona.objects.get(email=email)
            user = authenticate(request, username=email, password=password)
            if user is not None:
                # Si es admin, requerir verificación en dos pasos
                if persona.es_admin:
                    # Guardar los datos de usuario en la sesión temporalmente
                    request.session['temp_user_id'] = user.id
                    request.session['temp_user_email'] = email
                    
                    # Crear y enviar código de verificación
                    codigo = CodigoVerificacion.objects.create(
                        persona=persona,
                        fecha_expiracion=timezone.now() + timezone.timedelta(minutes=10)
                    )
                    
                    # Enviar el código por correo electrónico
                    subject = 'Código de verificación - Alquil.Ar'
                    message = f'''
                    Hola {persona.nombre},

                    Tu código de verificación es: {codigo.codigo}

                    Este código expirará en 10 minutos.

                    Si no solicitaste este código, alguien podría estar intentando acceder a tu cuenta.

                    Saludos,
                    El equipo de Alquil.Ar
                    '''
                    
                    send_mail(
                        subject,
                        message,
                        settings.DEFAULT_FROM_EMAIL,
                        [email],
                        fail_silently=False,
                    )
                    
                    messages.info(request, 'Por favor, ingresa el código de verificación enviado a tu correo.')
                    return redirect('persona:verificar_codigo')
                
                # Para usuarios no admin, proceder con el login normal
                if (((not persona.es_cliente) or persona.bloqueado_cliente) and ((not persona.es_empleado) or persona.bloqueado_empleado)):
                    return render(request, 'persona/login_unificado2.html', {
                        'error': 'Tu cuenta está suspendida. Contacta al administrador.'
                    })
                else:
                    login(request, user)
                    
                    # Verificar roles en orden específico
                    if persona.es_empleado:
                        messages.success(request, 'Has iniciado sesión como empleado')
                        return redirect('persona:inicio')
                    elif persona.es_cliente:
                        messages.success(request, 'Has iniciado sesión como cliente')
                        return redirect('persona:inicio')
                    else:
                        return render(request, 'persona/login_unificado2.html', {
                            'error': 'No tienes cuenta activa. Contacta al administrador.'
                        })
            else:
                 return render(request, 'persona/login_unificado2.html', {
                    'error': 'Email o contraseña incorrectos'
                 })
        except Persona.DoesNotExist:
            return render(request, 'persona/login_unificado2.html', {
                'error': 'Email o contraseña incorrectos'
            })

    # Si es GET, mostrar el formulario
    return render(request, 'persona/login_unificado2.html')

@empleado_requerido
def gestion(request):
    """Vista de gestión accesible para empleados y administradores"""
    # Obtener la persona del usuario actual
    try:
        usuario_persona = Persona.objects.get(email=request.user.email)
    except Persona.DoesNotExist:
        usuario_persona = None
    
    # Obtener lista de emails de empleados
    empleados_emails = list(Persona.objects.filter(es_empleado=True).values_list('email', flat=True))
    
    context = {
        'usuario_persona': usuario_persona,
        'empleados_emails': empleados_emails,
    }
    
    return render(request, 'persona/gestion.html', context)

@login_required
@user_passes_test(es_admin)
@csrf_protect
@ensure_csrf_cookie
@require_http_methods(["GET", "POST"])
def estadisticas(request):
    """Vista de estadísticas accesible solo para administradores"""
    return render(request, 'persona/estadisticas.html')

def empleados_processor(request):
    """Context processor que agrega datos de sesión y de usuario al contexto de todos los templates."""
    empleados_emails = list(Persona.objects.filter(es_empleado=True, bloqueado_empleado=False).values_list('email', flat=True))

    # Agregar la variable de sesión al contexto
    es_empleado_actuando_como_cliente = request.session.get('es_empleado_actuando_como_cliente', False)

    context = {
        'empleados_emails': empleados_emails,
        'es_empleado_actuando_como_cliente': es_empleado_actuando_como_cliente,
        'usuario_persona': None,
    }

    if request.user.is_authenticated:
        try:
            # Añadir el objeto Persona del usuario logueado para usarlo en los templates
            context['usuario_persona'] = Persona.objects.get(email=request.user.email)
        except Persona.DoesNotExist:
            pass # Mantiene usuario_persona como None si no hay perfil

    return context

@login_required
@csrf_protect
@ensure_csrf_cookie
@require_http_methods(["GET", "POST"])
def switch_back_to_employee(request):
    """Vista para volver a la cuenta de empleado desde la cuenta personal"""
    error = None
    
    # Verificar si el usuario está actuando como cliente pero es empleado
    if not request.session.get('es_empleado_actuando_como_cliente', False):
        messages.error(request, 'No estás en modo cliente.')
        return redirect('persona:inicio')
    
    try:
        persona = Persona.objects.get(email=request.user.email)
    except Persona.DoesNotExist:
        messages.error(request, 'No se encontró tu perfil.')
        return redirect('persona:inicio')
    
    if request.method == 'POST':
        # Verificar contraseña
        password = request.POST.get('password')
        if not password:
            error = 'Debes ingresar tu contraseña.'
        else:
            user = authenticate(request, username=request.user.email, password=password)
            if user is not None:
                # Eliminar la variable de sesión
                del request.session['es_empleado_actuando_como_cliente']
                messages.success(request, 'Has vuelto a tu cuenta de empleado exitosamente.')
                return redirect('persona:inicio')
            else:
                error = 'Contraseña incorrecta.'
    
    return render(request, 'persona/switch_back_to_employee.html', {
        'error': error,
        'persona': persona
    })

def logout_view(request):
    """Vista para cerrar sesión"""
    # Limpiar la variable de sesión que indica que el usuario es un empleado actuando como cliente
    if 'es_empleado_actuando_como_cliente' in request.session:
        del request.session['es_empleado_actuando_como_cliente']
    logout(request)
    messages.success(request, 'Cierre de sesión exitoso')
    return redirect('persona:inicio')

@login_required
@csrf_protect
@ensure_csrf_cookie
@require_http_methods(["GET", "POST"])
def modificar_datos_personales(request):
    """
    Vista para que los usuarios (clientes y empleados) modifiquen su nombre y apellido
    """
    from .forms import ModificarDatosPersonalesForm
    
    try:
        # Buscar el perfil por email del usuario
        persona = Persona.objects.get(email=request.user.email)
    except Persona.DoesNotExist:
        messages.error(request, "No se encontró el perfil asociado a tu email.")
        return redirect('persona:inicio')
    
    if request.method == 'POST':
        form = ModificarDatosPersonalesForm(request.POST, instance=persona)
        if form.is_valid():
            persona = form.save()
            
            # Actualizar también el usuario de Django
            request.user.first_name = persona.nombre
            request.user.last_name = persona.apellido
            request.user.save()
            
            messages.success(request, 'Datos actualizados correctamente.')
            return redirect('persona:inicio')
    else:
        form = ModificarDatosPersonalesForm(instance=persona)
    
    return render(request, 'persona/modificar_datos_personales.html', {
        'form': form,
        'persona': persona
    })

@csrf_exempt
def webhook_mercadopago(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            print(f"=== WEBHOOK MERCADOPAGO RECIBIDO ===")
            print(f"Data: {data}")
            
            # Obtener el ID del pago
            if data.get('type') == 'payment' and data.get('data', {}).get('id'):
                payment_id = data['data']['id']
                
                # Configurar SDK
                sdk = mercadopago.SDK(settings.MERCADOPAGO_ACCESS_TOKEN)
                
                # Obtener información del pago
                payment_info = sdk.payment().get(payment_id)
                
                if payment_info['status'] == 200:
                    payment = payment_info['response']
                    status = payment.get('status')
                    external_reference = payment.get('external_reference')
                    
                    print(f"Payment ID: {payment_id}")
                    print(f"Status: {status}")
                    print(f"External Reference: {external_reference}")
                    
                    if status == 'approved' and external_reference:
                        # Parsear external_reference
                        # Formato: "maquina_id|persona_id|fecha_inicio|fecha_fin|metodo_pago|total"
                        try:
                            datos = external_reference.split('|')
                            if len(datos) == 6:
                                from datetime import datetime
                                
                                maquina_id, persona_id, fecha_inicio_str, fecha_fin_str, metodo_pago, total = datos
                                
                                # Convertir fechas
                                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
                                fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
                                
                                # Obtener objetos
                                maquina_base = MaquinaBase.objects.get(id=maquina_id)
                                persona = Persona.objects.get(id=persona_id)
                                
                                # Verificar que no existe ya un alquiler para estos datos
                                alquiler_existente = Alquiler.objects.filter(
                                    persona=persona,
                                    maquina_base=maquina_base,
                                    fecha_inicio=fecha_inicio,
                                    fecha_fin=fecha_fin
                                ).first()
                                
                                if not alquiler_existente:
                                    # Verificar disponibilidad nuevamente
                                    if Alquiler.verificar_disponibilidad(maquina_base, fecha_inicio, fecha_fin):
                                        # Verificar que el cliente no tenga otro alquiler activo
                                        alquileres_activos = Alquiler.objects.filter(
                                            persona=persona,
                                            estado__in=['reservado', 'en_curso']
                                        )
                                        
                                        if not alquileres_activos.exists():
                                            # Crear el alquiler
                                            alquiler = Alquiler.objects.create(
                                                maquina_base=maquina_base,
                                                persona=persona,
                                                fecha_inicio=fecha_inicio,
                                                fecha_fin=fecha_fin,
                                                metodo_pago=metodo_pago,
                                                estado='reservado',
                                                monto_total=float(total),
                                                preference_id=payment.get('order', {}).get('id')
                                            )
                                            
                                            print(f"Alquiler creado: {alquiler.numero}")
                                            
                                            # Enviar mail al cliente
                                            try:
                                                send_mail(
                                                    'Alquiler confirmado - ALQUIL.AR',
                                                    f'Alquiler confirmado, gracias por alquilar con ALQUIL.AR.\n'
                                                    f'Aquí te dejamos tu número de retiro: {alquiler.numero}',
                                                    settings.DEFAULT_FROM_EMAIL,
                                                    [persona.email],
                                                    fail_silently=False,
                                                )
                                                print(f"Email enviado exitosamente a: {persona.email}")
                                            except Exception as e:
                                                print(f"Error al enviar email: {str(e)}")
                                            
                                            print(f"=== ALQUILER CREADO VIA WEBHOOK ===")
                                            print(f"ID: {alquiler.id}")
                                            print(f"Número: {alquiler.numero}")
                                            
                                            # Actualizar estado de la máquina si es necesaria
                                            if alquiler.unidad:
                                                maquina = Maquina.objects.filter(
                                                    patente=alquiler.unidad.patente
                                                ).first()
                                                if maquina:
                                                    maquina.estado = 'alquilada'
                                                    maquina.save()
                                                    print(f"Máquina actualizada: {maquina.id} - Estado: {maquina.estado}")
                                else:
                                    print(f"Alquiler ya existe: {alquiler_existente.numero}")
                                    
                        except Exception as e:
                            print(f"Error al procesar webhook: {str(e)}")
                            
            return HttpResponse(status=200)
            
        except Exception as e:
            print(f"Error en webhook: {str(e)}")
            return HttpResponse(status=500)
            
    return HttpResponse(status=405)

@login_required
def checkout(request, alquiler_id):
    alquiler = get_object_or_404(Alquiler, id=alquiler_id)
    status = request.GET.get('status')
    
    # Verificar que el alquiler pertenece al usuario actual
    if alquiler.persona.email != request.user.email:
        messages.error(request, "No tienes permiso para acceder a este alquiler.")
        return redirect('persona:mis_alquileres')
    
    if status == 'approved':
        # Actualizar el estado del alquiler
        alquiler.estado = 'confirmado'
        alquiler.save()
        
        # Marcar la máquina como alquilada
        maquina = alquiler.maquina
        maquina.estado = 'alquilada'
        maquina.save()
        
        messages.success(request, "¡Pago exitoso! Tu alquiler ha sido confirmado.")
    elif status == 'pending':
        alquiler.estado = 'pendiente'
        alquiler.save()
        messages.info(request, "El pago está pendiente de confirmación.")
    else:
        alquiler.estado = 'fallido'
        alquiler.save()
        messages.error(request, "El pago no pudo ser procesado. Por favor, intenta nuevamente.")
    
    return redirect('persona:mis_alquileres')

def procesar_alquileres_vencidos_automatico():
    """Función para procesar automáticamente alquileres vencidos"""
    from datetime import date
    
    # Obtener la fecha actual
    fecha_actual = date.today()
    
    # Buscar alquileres vencidos que no han sido devueltos
    alquileres_vencidos = Alquiler.objects.filter(
        estado='en_curso',  # Solo alquileres en curso
        fecha_fin__lt=fecha_actual  # Fecha de fin menor a la fecha actual
    )
    
    # También revisar alquileres reservados vencidos (que nunca se retiraron)
    alquileres_reservados_vencidos = Alquiler.objects.filter(
        estado='reservado',  # Alquileres reservados
        fecha_inicio__lt=fecha_actual  # Fecha de inicio menor a la fecha actual
    )
    
    procesados = 0
    
    # Procesar alquileres en curso vencidos
    for alquiler in alquileres_vencidos:
        try:
            # Cambiar estado del alquiler a "adeudado"
            alquiler.estado = 'adeudado'
            alquiler.save()
            
            # Poner la máquina en mantenimiento
            if alquiler.unidad:
                alquiler.unidad.estado = 'adeudado'
                alquiler.unidad.save()
            
            procesados += 1
            
        except Exception as e:
            # Log del error silencioso para no interrumpir la vista
            print(f"Error procesando alquiler vencido {alquiler.numero}: {str(e)}")
    
    # Procesar alquileres reservados que nunca se retiraron (cancelar automáticamente)
    for alquiler in alquileres_reservados_vencidos:
        try:
            # Cambiar estado a cancelado en lugar de adeudado
            alquiler.estado = 'cancelado'
            alquiler.fecha_cancelacion = fecha_actual
            alquiler.cancelado_por_empleado = True
            alquiler.observaciones_cancelacion = "Cancelado automáticamente por no retirar en fecha de inicio"
            alquiler.porcentaje_reembolso = 0  # Sin reembolso por no retirar
            alquiler.monto_reembolso = 0
            alquiler.save()
            
            procesados += 1
            
        except Exception as e:
            print(f"Error procesando alquiler reservado vencido {alquiler.numero}: {str(e)}")
    
    return procesados

@empleado_requerido
def lista_alquileres(request):
    """Vista completa de gestión de alquileres para empleados y admins"""
    # Procesar automáticamente alquileres vencidos
    alquileres_procesados = procesar_alquileres_vencidos_automatico()
    if alquileres_procesados > 0:
        print(f"Se procesaron automáticamente {alquileres_procesados} alquileres vencidos")
    
    # Obtener todos los alquileres inicialmente
    alquileres = Alquiler.objects.select_related(
        'maquina_base', 'unidad', 'persona', 'unidad__sucursal'
    ).all()
    
    # Aplicar filtros
    estado_filtro = request.GET.get('estado')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    cliente_filtro = request.GET.get('cliente')
    sucursal_filtro = request.GET.get('sucursal')
    
    if estado_filtro:
        alquileres = alquileres.filter(estado=estado_filtro)
    
    if fecha_desde:
        try:
            from datetime import datetime
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            alquileres = alquileres.filter(fecha_inicio__gte=fecha_desde_obj)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            from datetime import datetime
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            alquileres = alquileres.filter(fecha_fin__lte=fecha_hasta_obj)
        except ValueError:
            pass
    
    if cliente_filtro:
        alquileres = alquileres.filter(
            Q(persona__nombre__icontains=cliente_filtro) |
            Q(persona__apellido__icontains=cliente_filtro) |
            Q(persona__email__icontains=cliente_filtro) |
            Q(persona__dni__icontains=cliente_filtro)
        )
    
    if sucursal_filtro:
        alquileres = alquileres.filter(unidad__sucursal_id=sucursal_filtro)
    
    # Comprobar si se ha aplicado algún filtro
    filtros_aplicados = bool(estado_filtro or fecha_desde or fecha_hasta or cliente_filtro or sucursal_filtro)

    # Ordenar por fecha de creación más reciente
    alquileres = alquileres.order_by('-fecha_creacion')
    
    # Calcular estadísticas
    stats = Alquiler.objects.aggregate(
        reservados=Count('id', filter=Q(estado='reservado')),
        en_curso=Count('id', filter=Q(estado='en_curso')),
        finalizados=Count('id', filter=Q(estado='finalizado')),
        cancelados=Count('id', filter=Q(estado='cancelado')),
        adeudados=Count('id', filter=Q(estado='adeudado'))
    )
    
    # Obtener sucursales para el filtro
    sucursales = Sucursal.objects.all().order_by('direccion')
    
    # Paginación
    paginator = Paginator(alquileres, 25)  # 25 alquileres por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Verificar si se solicita exportación
    if request.GET.get('export') == 'xlsx':
        return exportar_alquileres_xlsx(alquileres)
    
    # Mostrar mensaje si no hay resultados y se aplicaron filtros
    mensaje_sin_resultados = None
    if not page_obj.object_list and filtros_aplicados:
        mensaje_sin_resultados = "No se encontraron alquileres con los filtros aplicados."
    
    context = {
        'alquileres': page_obj,
        'stats': stats,
        'sucursales': sucursales,
        'mensaje_sin_resultados': mensaje_sin_resultados,
        'filtros': {
            'estado': estado_filtro,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'cliente': cliente_filtro,
            'sucursal': sucursal_filtro,
        }
    }
    
    return render(request, 'persona/lista_alquileres.html', context)

@empleado_requerido
@csrf_protect
@require_http_methods(["POST"])
def iniciar_alquiler(request):
    """Vista para iniciar un alquiler verificando el código de retiro"""
    try:
        alquiler_id = request.POST.get('alquiler_id')
        codigo_cliente = request.POST.get('codigo_cliente', '').strip()
        
        if not alquiler_id or not codigo_cliente:
            return JsonResponse({
                'success': False,
                'error': 'Faltan datos requeridos.'
            })
        
        # Buscar el alquiler
        try:
            alquiler = Alquiler.objects.get(id=alquiler_id)
        except Alquiler.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Alquiler no encontrado.'
            })
        
        # Verificar que el alquiler esté en estado 'reservado'
        if alquiler.estado != 'reservado':
            return JsonResponse({
                'success': False,
                'error': f'El alquiler está en estado {alquiler.get_estado_display()}. Solo se pueden iniciar alquileres reservados.'
            })
        
        # Verificar el código de retiro
        if alquiler.codigo_retiro != codigo_cliente:
            return JsonResponse({
                'success': False,
                'error': 'El código ingresado no coincide con el código de retiro del alquiler. Verifica el código con el cliente.'
            })
        
        # Validar que la fecha de inicio sea igual a la fecha actual
        from datetime import date
        fecha_actual = date.today()
        if alquiler.fecha_inicio != fecha_actual:
            if alquiler.fecha_inicio > fecha_actual:
                return JsonResponse({
                    'success': False,
                    'error': f'No se puede retirar el alquiler aún. La fecha de inicio es {alquiler.fecha_inicio.strftime("%d/%m/%Y")} y hoy es {fecha_actual.strftime("%d/%m/%Y")}.'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': f'El alquiler venció. La fecha de inicio era {alquiler.fecha_inicio.strftime("%d/%m/%Y")} y hoy es {fecha_actual.strftime("%d/%m/%Y")}. Contacta con un administrador.'
                })
        
        # Cambiar estado a 'en_curso'
        alquiler.estado = 'en_curso'
        alquiler.save()
        
        # Opcional: Enviar email de confirmación de inicio
        try:
            from maquinas.utils import enviar_email_inicio_alquiler
            enviar_email_inicio_alquiler(alquiler)
        except Exception as e:
            print(f"Error enviando email de inicio: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'message': f'Alquiler {alquiler.numero} iniciado correctamente.'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error interno: {str(e)}'
        })

@empleado_requerido
@csrf_protect
@require_http_methods(["POST"])
def finalizar_alquiler(request):
    """Vista para finalizar un alquiler con verificación de código y calificación del cliente"""
    try:
        from maquinas.models import CalificacionCliente
        from maquinas.utils import enviar_email_finalizacion_alquiler
        
        alquiler_id = request.POST.get('alquiler_id')
        codigo_cliente = request.POST.get('codigo_cliente', '').strip()
        calificacion = request.POST.get('calificacion')
        observaciones = request.POST.get('observaciones', '').strip()
        # Interpretar el checkbox de máquina dañada
        maquina_danada_raw = request.POST.get('maquina_danada')
        maquina_danada = maquina_danada_raw in ['true', 'on', '1', True]
        
        # Debug logs removidos para producción
        
        # Validar datos requeridos
        if not alquiler_id or not codigo_cliente or not calificacion:
            return JsonResponse({
                'success': False,
                'error': 'Faltan datos requeridos.'
            })
        
        # Validar calificación
        try:
            calificacion = int(calificacion)
            if calificacion < 1 or calificacion > 5:
                raise ValueError("Calificación fuera de rango")
        except (ValueError, TypeError):
            return JsonResponse({
                'success': False,
                'error': 'La calificación debe ser un número entre 1 y 5.'
            })
        
        # Buscar el alquiler
        try:
            alquiler = Alquiler.objects.get(id=alquiler_id)
        except Alquiler.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Alquiler no encontrado.'
            })
        
        # Verificar que el alquiler esté en estado 'en_curso' o 'adeudado'
        if alquiler.estado not in ['en_curso', 'adeudado']:
            return JsonResponse({
                'success': False,
                'error': f'No es posible devolver en este estado. El alquiler está en estado {alquiler.get_estado_display()}.'
            })
        
        # Verificar el código de retiro
        if alquiler.codigo_retiro != codigo_cliente:
            return JsonResponse({
                'success': False,
                'error': 'DNI o código de reserva inválido.'
            })
        
        # Cambiar estado del alquiler a 'finalizado'
        alquiler.estado = 'finalizado'
        alquiler.save()
        
        # Actualizar estado de la máquina/unidad
        if alquiler.unidad:
            if maquina_danada:
                alquiler.unidad.estado = 'mantenimiento'
                mensaje_maquina = "Máquina marcada en mantenimiento."
            else:
                alquiler.unidad.estado = 'disponible'
                mensaje_maquina = "Máquina marcada como disponible."
            alquiler.unidad.save()
        else:
            mensaje_maquina = "Sin unidad asignada."
        
        # Crear calificación del cliente
        CalificacionCliente.objects.create(
            alquiler=alquiler,
            cliente=alquiler.persona,
            empleado=request.user,
            calificacion=calificacion,
            observaciones=observaciones
        )
        
        # Enviar email de finalización (opcional)
        try:
            enviar_email_finalizacion_alquiler(alquiler)
        except Exception as e:
            print(f"Error enviando email de finalización: {str(e)}")
        
        # Mensaje de éxito
        if alquiler.estado == 'adeudado':
            mensaje = f"Devolución tardía registrada. {mensaje_maquina} Cliente calificado con {calificacion} estrellas."
        else:
            mensaje = f"Devolución registrada. {mensaje_maquina} Cliente calificado con {calificacion} estrellas."
        
        return JsonResponse({
            'success': True,
            'message': mensaje
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error interno: {str(e)}'
        })

def exportar_alquileres_xlsx(alquileres):
    """Exportar alquileres a Excel"""
    try:
        import openpyxl
        from django.http import HttpResponse
        from datetime import datetime
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Alquileres"
        
        # Encabezados
        headers = [
            'Número', 'Cliente', 'Email', 'Teléfono', 'Máquina', 'Marca', 'Modelo',
            'Unidad', 'Sucursal', 'Fecha Inicio', 'Fecha Fin', 'Días', 'Estado',
            'Método Pago', 'Monto', 'Código Retiro', 'Fecha Creación'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Datos
        for row, alquiler in enumerate(alquileres, 2):
            ws.cell(row=row, column=1, value=alquiler.numero)
            ws.cell(row=row, column=2, value=f"{alquiler.persona.nombre} {alquiler.persona.apellido}" if alquiler.persona else "N/A")
            ws.cell(row=row, column=3, value=alquiler.persona.email if alquiler.persona else "N/A")
            ws.cell(row=row, column=4, value=alquiler.persona.telefono if alquiler.persona else "N/A")
            ws.cell(row=row, column=5, value=alquiler.maquina_base.nombre)
            ws.cell(row=row, column=6, value=alquiler.maquina_base.get_marca_display())
            ws.cell(row=row, column=7, value=alquiler.maquina_base.modelo)
            ws.cell(row=row, column=8, value=alquiler.unidad.patente if alquiler.unidad else "Sin asignar")
            ws.cell(row=row, column=9, value=alquiler.unidad.sucursal.direccion if alquiler.unidad else "N/A")
            ws.cell(row=row, column=10, value=alquiler.fecha_inicio.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=11, value=alquiler.fecha_fin.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=12, value=alquiler.cantidad_dias)
            ws.cell(row=row, column=13, value=alquiler.get_estado_display())
            ws.cell(row=row, column=14, value=alquiler.get_metodo_pago_display())
            ws.cell(row=row, column=15, value=float(alquiler.monto_total) if alquiler.monto_total else 0)
            ws.cell(row=row, column=16, value=alquiler.codigo_retiro)
            ws.cell(row=row, column=17, value=alquiler.fecha_creacion.strftime('%d/%m/%Y %H:%M'))
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="alquileres_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'No se pudo exportar a Excel. Instala openpyxl.')
        return redirect('persona:lista_alquileres')

def recuperar_password(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            # Buscar la persona por email
            persona = Persona.objects.get(email=email)
            
            # Buscar todos los usuarios asociados
            users = User.objects.filter(email=email)
            if users.exists():
                # Generar nueva contraseña aleatoria
                nueva_password = generar_password_random()
                
                # Actualizar la contraseña de todos los usuarios con ese email
                for user in users:
                    user.set_password(nueva_password)
                    user.save()
                
                # Enviar email con la nueva contraseña
                send_mail(
                    'Alquil.Ar - Tu nueva contraseña',
                    f'Hola {persona.nombre},\n\n'
                    f'Has solicitado una nueva contraseña.\n\n'
                    f'Tu nueva contraseña es: {nueva_password}\n\n'
                    f'Por favor, cambia tu contraseña la próxima vez que inicies sesión.\n\n'
                    f'Saludos,\n'
                    f'Equipo Alquil.Ar',
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=False,
                )
            
        except (Persona.DoesNotExist, User.DoesNotExist, Exception):
            # Si el email no existe o hay algún error, simplemente continuamos
            pass
        
        # Siempre mostramos mensaje de éxito
        messages.success(request, 'Se ha enviado una nueva contraseña a tu email.')
        return redirect('persona:login_unificado2')
    
    return render(request, 'persona/recuperar_password.html')

def mapa_sucursales(request):
    try:
        sucursales = Sucursal.objects.all()
        
        # Validar que las coordenadas sean válidas
        for sucursal in sucursales:
            if not (-90 <= sucursal.latitud <= 90) or not (-180 <= sucursal.longitud <= 180):
                raise ValueError(f'Coordenadas inválidas en sucursal {sucursal.id_sucursal}')
        
        return render(request, 'mapa_sucursales.html', {
            'sucursales': sucursales,
            'centro_mapa': [-34.9214, -57.9544]  # Coordenadas de La Plata
        })
    except ValueError as e:
        messages.error(request, f'Error en las coordenadas: {str(e)}')
        return redirect('persona:inicio')
    except Exception as e:
        messages.error(request, 'Ocurrió un error al cargar el mapa de sucursales')
        return redirect('persona:inicio')

@login_required
def pago_exitoso(request):
    """Página de pago exitoso con confetis"""
    # Obtener datos del pago si están disponibles
    external_reference = request.GET.get('external_reference')
    payment_id = request.GET.get('payment_id')
    
    alquiler = None
    
    print(f"=== PAGO EXITOSO - PROCESANDO RETORNO ===")
    print(f"External reference: {external_reference}")
    print(f"Payment ID: {payment_id}")
    
    # Procesar el pago si tenemos external_reference (usuario viene de MercadoPago)
    if external_reference:
        try:
            datos = external_reference.split('|')
            if len(datos) == 6:
                from datetime import datetime
                
                maquina_id, persona_id, fecha_inicio_str, fecha_fin_str, metodo_pago, total = datos
                
                # Convertir fechas
                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
                fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
                
                # Obtener objetos
                maquina_base = MaquinaBase.objects.get(id=maquina_id)
                persona = Persona.objects.get(id=persona_id)
                
                # Verificar que el usuario actual es el mismo que hizo el pago
                try:
                    persona_actual = Persona.objects.get(email=request.user.email)
                    if persona.id != persona_actual.id:
                        messages.error(request, "Error de validación de usuario.")
                        return redirect('persona:inicio')
                except Persona.DoesNotExist:
                    messages.error(request, "Usuario no encontrado.")
                    return redirect('persona:inicio')
                
                # IDEMPOTENCIA: Verificar que no existe ya un alquiler ACTIVO para esta referencia externa
                alquiler_por_referencia = Alquiler.objects.filter(
                    external_reference=external_reference,
                    estado__in=['reservado', 'en_curso']  # Solo alquileres realmente activos
                ).first()
                
                if alquiler_por_referencia:
                    print(f"[PAGO_EXITOSO] Alquiler ya existe por referencia: {alquiler_por_referencia.numero}")
                    alquiler = alquiler_por_referencia
                else:
                    # Verificar que no existe ya un alquiler ACTIVO para estos datos
                    alquiler_existente = Alquiler.objects.filter(
                        persona=persona,
                        maquina_base=maquina_base,
                        fecha_inicio=fecha_inicio,
                        fecha_fin=fecha_fin,
                        estado__in=['reservado', 'en_curso']  # Solo alquileres realmente activos
                    ).first()
                    
                    if alquiler_existente:
                        print(f"[PAGO_EXITOSO] Alquiler ya existe por datos: {alquiler_existente.numero}")
                        # Actualizar la referencia externa si no la tiene
                        if not alquiler_existente.external_reference:
                            alquiler_existente.external_reference = external_reference
                            alquiler_existente.save()
                        alquiler = alquiler_existente
                    else:
                        # Verificar disponibilidad nuevamente
                        if Alquiler.verificar_disponibilidad(maquina_base, fecha_inicio, fecha_fin):
                            # Verificar que el cliente no tenga otro alquiler activo
                            alquileres_activos = Alquiler.objects.filter(
                                persona=persona,
                                estado__in=['reservado', 'en_curso']
                            )
                            
                            if not alquileres_activos.exists():
                                # Crear el alquiler
                                alquiler = Alquiler.objects.create(
                                    maquina_base=maquina_base,
                                    persona=persona,
                                    fecha_inicio=fecha_inicio,
                                    fecha_fin=fecha_fin,
                                    metodo_pago=metodo_pago,
                                    estado='reservado',
                                    monto_total=float(total),
                                    external_reference=external_reference
                                )
                                
                                print(f"[PAGO_EXITOSO] Alquiler creado via fallback: {alquiler.numero}")
                                
                                # Enviar email con PDF al cliente
                                try:
                                    print(f"[PAGO_EXITOSO] Intentando enviar email...")
                                    resultado_email = enviar_email_alquiler_simple(alquiler)
                                    if resultado_email:
                                        print(f"[PAGO_EXITOSO] Email enviado correctamente")
                                    else:
                                        print(f"[PAGO_EXITOSO] Falló el envío de email")
                                except Exception as e:
                                    print(f"[PAGO_EXITOSO] Error al enviar email: {str(e)}")
                            else:
                                print(f"[PAGO_EXITOSO] Cliente ya tiene alquiler activo")
                                messages.warning(request, 'Ya tienes un alquiler activo.')
                        else:
                            print(f"[PAGO_EXITOSO] No hay disponibilidad")
                            messages.error(request, 'No hay unidades disponibles para las fechas seleccionadas.')
                    
        except Exception as e:
            print(f"Error al procesar retorno de pago: {str(e)}")
            messages.error(request, 'Hubo un error al procesar tu pago. Contacta al soporte.')
    
    # Si no encontramos el alquiler específico, buscar el más reciente del usuario
    if not alquiler:
        try:
            persona = Persona.objects.get(email=request.user.email)
            alquiler = Alquiler.objects.filter(
                persona=persona
            ).order_by('-fecha_creacion').first()
        except:
            pass
    
    return render(request, 'persona/pago_exitoso.html', {
        'alquiler': alquiler
    })

@login_required 
def pago_fallido(request):
    """Página de pago fallido"""
    return render(request, 'persona/pago_fallido.html')

@login_required
def pago_pendiente(request):
    """Página para notificar que el pago está pendiente"""
    return render(request, 'persona/pago_pendiente.html')

@empleado_requerido
def lista_reembolsos(request):
    """Vista para gestionar reembolsos pendientes y pagados"""
    from maquinas.models import Reembolso
    from django.db.models import Q
    from django.core.paginator import Paginator
    
    # Obtener todos los reembolsos
    reembolsos = Reembolso.objects.select_related(
        'alquiler', 'alquiler__persona', 'alquiler__maquina_base', 
        'empleado_que_marco_pagado'
    ).all()
    
    # Aplicar filtros
    estado_filtro = request.GET.get('estado')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    cliente_filtro = request.GET.get('cliente')
    
    # Filtrar por estado (solo pendiente y pagado)
    if estado_filtro and estado_filtro in ['pendiente', 'pagado']:
        reembolsos = reembolsos.filter(estado=estado_filtro)
    
    # Filtrar por fechas
    if fecha_desde:
        try:
            from datetime import datetime
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            reembolsos = reembolsos.filter(fecha_creacion__date__gte=fecha_desde_obj)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            from datetime import datetime
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            reembolsos = reembolsos.filter(fecha_creacion__date__lte=fecha_hasta_obj)
        except ValueError:
            pass
    
    # Filtrar por cliente
    if cliente_filtro:
        reembolsos = reembolsos.filter(
            Q(alquiler__persona__nombre__icontains=cliente_filtro) |
            Q(alquiler__persona__apellido__icontains=cliente_filtro) |
            Q(alquiler__persona__email__icontains=cliente_filtro) |
            Q(alquiler__persona__dni__icontains=cliente_filtro)
        )
    
    # Ordenar por fecha de creación más reciente
    reembolsos = reembolsos.order_by('-fecha_creacion')
    
    # Calcular estadísticas
    from django.db.models import Count, Sum
    stats = Reembolso.objects.aggregate(
        pendientes=Count('id', filter=Q(estado='pendiente')),
        pagados=Count('id', filter=Q(estado='pagado')),
        total_pendiente=Sum('monto', filter=Q(estado='pendiente')),
        total_pagado=Sum('monto', filter=Q(estado='pagado'))
    )
    
    # Paginación
    paginator = Paginator(reembolsos, 25)  # 25 reembolsos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'reembolsos': page_obj,
        'stats': stats,
        'filtros': {
            'estado': estado_filtro,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'cliente': cliente_filtro,
        }
    }
    
    return render(request, 'persona/lista_reembolsos.html', context)

@empleado_requerido
def marcar_reembolso_pagado(request, reembolso_id):
    """Vista para marcar un reembolso como pagado"""
    from maquinas.models import Reembolso
    
    reembolso = get_object_or_404(Reembolso, id=reembolso_id)
    
    if request.method == 'POST':
        try:
            if reembolso.estado != 'pendiente':
                messages.error(request, f'El reembolso ya está {reembolso.get_estado_display().lower()}')
                return redirect('persona:lista_reembolsos')
            
            observaciones = request.POST.get('observaciones', '')
            reembolso.marcar_como_pagado(empleado=request.user, observaciones=observaciones)
            
            messages.success(request, 
                f'Reembolso del alquiler {reembolso.alquiler.numero} marcado como pagado exitosamente. '
                f'Monto: ${reembolso.monto}')
            
        except Exception as e:
            messages.error(request, f'Error al marcar el reembolso como pagado: {str(e)}')
        
        return redirect('persona:lista_reembolsos')
    
    return render(request, 'persona/marcar_reembolso_pagado.html', {
        'reembolso': reembolso
    })

def buscar_clientes_json(request):
    """API endpoint para buscar clientes dinámicamente - SOLO POR EMAIL Y DNI - SIN AUTENTICACIÓN"""
    from django.http import JsonResponse
    from django.db.models import Q
    
    query = request.GET.get('q', '').strip()
    
    print(f"🔍 API BÚSQUEDA: '{query}' (longitud: {len(query)})")
    print(f"📊 Usuario: {request.user} - Autenticado: {request.user.is_authenticated}")
    
    if len(query) < 2:  # Mínimo 2 caracteres para buscar
        print("[INFO] Búsqueda muy corta, devolviendo lista vacía")
        return JsonResponse({'clientes': []})
    
    # Buscar SOLO por email y DNI en todas las personas (excluir empleados)
    print(f"📊 Total personas en BD: {Persona.objects.count()}")
    
    clientes = Persona.objects.filter(
        Q(email__icontains=query) | Q(dni__icontains=query),
        es_empleado=False  # Excluir empleados
    ).order_by('apellido', 'nombre')[:10]  # Limitar a 10 resultados
    
    print(f"🔍 Total encontrados: {clientes.count()}")
    
    resultados = []
    for cliente in clientes:
        display_text = f"{cliente.apellido}, {cliente.nombre} - {cliente.email}"
        if cliente.dni:
            display_text += f" (DNI: {cliente.dni})"
            
        resultados.append({
            'id': cliente.id,
            'nombre': cliente.nombre,
            'apellido': cliente.apellido,
            'nombre_completo': f"{cliente.nombre} {cliente.apellido}",
            'email': cliente.email,
            'dni': cliente.dni if cliente.dni else '',
            'display_text': display_text,
            'calificacion_promedio': float(cliente.calificacion_promedio)
        })
        print(f"   📄 Cliente: {cliente.email} - DNI: {cliente.dni}")
    
    print(f"📤 Devolviendo {len(resultados)} resultados")
    return JsonResponse({'clientes': resultados})

@login_required
def obtener_datos_cliente_actual(request):
    """Vista para obtener los datos del cliente actual para cálculo de recargo"""
    try:
        persona = Persona.objects.get(email=request.user.email)
        return JsonResponse({
            'success': True,
            'cliente': {
                'id': persona.id,
                'nombre': persona.nombre,
                'apellido': persona.apellido,
                'email': persona.email,
                'calificacion_promedio': float(persona.calificacion_promedio),
                'tiene_recargo': persona.tiene_recargo()
            }
        })
    except Persona.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Cliente no encontrado'
        })

def enviar_codigo_verificacion(request):
    if not request.user.is_authenticated:
        messages.error(request, 'Debes iniciar sesión primero.')
        return redirect('login')

    # Crear nuevo código de verificación
    codigo = CodigoVerificacion.objects.create(
        persona=request.user,
        fecha_expiracion=timezone.now() + timezone.timedelta(minutes=10)
    )

    # Enviar el código por correo electrónico
    subject = 'Código de verificación - Alquil.Ar'
    message = f'''
    Hola {request.user.nombre},

    Tu código de verificación es: {codigo.codigo}

    Este código expirará en 10 minutos.

    Si no solicitaste este código, puedes ignorar este mensaje.

    Saludos,
    El equipo de Alquil.Ar
    '''
    
    send_mail(
        subject,
        message,
        settings.DEFAULT_FROM_EMAIL,
        [request.user.email],
        fail_silently=False,
    )

    messages.success(request, 'Se ha enviado un código de verificación a tu correo electrónico.')
    return redirect('verificar_codigo')

def verificar_codigo(request):
    # Verificar si hay un usuario temporal en la sesión
    temp_user_id = request.session.get('temp_user_id')
    temp_user_email = request.session.get('temp_user_email')
    
    if not temp_user_id or not temp_user_email:
        messages.error(request, 'Sesión inválida. Por favor, inicia sesión nuevamente.')
        return redirect('persona:login_unificado2')

    try:
        persona = Persona.objects.get(email=temp_user_email)
    except Persona.DoesNotExist:
        messages.error(request, 'Usuario no encontrado. Por favor, inicia sesión nuevamente.')
        return redirect('persona:login_unificado2')

    if request.method == 'POST':
        codigo_ingresado = request.POST.get('codigo')
        
        # Buscar el código más reciente no usado y no expirado
        try:
            codigo = CodigoVerificacion.objects.filter(
                persona=persona,
                usado=False,
                fecha_expiracion__gt=timezone.now()
            ).latest('fecha_creacion')
            
            if codigo.codigo == codigo_ingresado:
                codigo.usado = True
                codigo.save()
                
                # Obtener el usuario y hacer login
                user = User.objects.get(id=temp_user_id)
                
                # Si es admin, asignar permisos completos
                if persona.es_admin:
                    user.is_staff = True
                    user.is_superuser = True
                    user.save()
                
                login(request, user)
                
                # Limpiar datos temporales de la sesión de forma segura
                request.session.pop('temp_user_id', None)
                request.session.pop('temp_user_email', None)
                
                messages.success(request, '¡Verificación exitosa! Has iniciado sesión como administrador con todos los permisos.')
                return redirect('persona:inicio')
            else:
                messages.error(request, 'Código incorrecto. Por favor, intenta nuevamente.')
        except CodigoVerificacion.DoesNotExist:
            messages.error(request, 'El código ha expirado o no existe. Por favor, inicia sesión nuevamente.')
            return redirect('persona:login_unificado2')

    return render(request, 'persona/verificar_codigo.html')

@empleado_requerido
def lista_clientes(request):
    queryset = Persona.objects.filter(es_cliente=True, es_admin=False).order_by('-fecha_registro')

    filtros = {
        'nombre': request.GET.get('nombre', ''),
        'dni': request.GET.get('dni', ''),
        'email': request.GET.get('email', ''),
        'estado': request.GET.get('estado', ''),
        'fecha_desde': request.GET.get('fecha_desde', ''),
        'fecha_hasta': request.GET.get('fecha_hasta', ''),
        'calificacion_desde': request.GET.get('calificacion_desde', ''),
        'calificacion_hasta': request.GET.get('calificacion_hasta', ''),
    }

    if filtros['nombre']:
        queryset = queryset.filter(
            Q(nombre__icontains=filtros['nombre']) | Q(apellido__icontains=filtros['nombre'])
        )
    if filtros['dni']:
        queryset = queryset.filter(dni__icontains=filtros['dni'])
    if filtros['email']:
        queryset = queryset.filter(email__icontains=filtros['email'])
    if filtros['estado']:
        if filtros['estado'] == 'activo':
            queryset = queryset.filter(bloqueado_cliente=False)
        elif filtros['estado'] == 'bloqueado':
            queryset = queryset.filter(bloqueado_cliente=True)
    if filtros['fecha_desde']:
        queryset = queryset.filter(fecha_registro__gte=filtros['fecha_desde'])
    if filtros['fecha_hasta']:
        # Añadimos un día para que la búsqueda sea inclusiva
        from datetime import datetime, timedelta
        fecha_hasta_dt = datetime.strptime(filtros['fecha_hasta'], '%Y-%m-%d') + timedelta(days=1)
        queryset = queryset.filter(fecha_registro__lt=fecha_hasta_dt)
    
    # El filtrado por calificación se hace usando el campo del modelo Persona
    if filtros['calificacion_desde']:
        queryset = queryset.filter(calificacion_promedio__gte=filtros['calificacion_desde'])
    if filtros['calificacion_hasta']:
        queryset = queryset.filter(calificacion_promedio__lte=filtros['calificacion_hasta'])

    # Comprobar si se ha aplicado algún filtro
    filtros_aplicados = any(filtros.values())

    # Exportar a Excel si se solicita
    if request.GET.get('export') == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes"

        # Encabezados
        headers = [
            'Nombre', 'Apellido', 'Email', 'DNI', 
            'Fecha Nacimiento', 'Fecha Registro', 'Calificación Promedio', 'Estado'
        ]
        ws.append(headers)

        # Estilos para cabecera
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Datos
        for cliente in queryset:
            estado = "Bloqueado" if cliente.bloqueado_cliente else "Activo"
            calificacion = cliente.calificacion_promedio if cliente.calificacion_promedio is not None else 5.0
            
            # Formatear fecha de nacimiento si existe
            fecha_nacimiento_str = cliente.fecha_nacimiento.strftime('%d/%m/%Y') if cliente.fecha_nacimiento else 'N/A'

            ws.append([
                cliente.nombre,
                cliente.apellido,
                cliente.email,
                cliente.dni or 'N/A',
                fecha_nacimiento_str,
                cliente.fecha_registro.strftime('%d/%m/%Y %H:%M'),
                round(calificacion, 2),
                estado
            ])

        # Ajustar ancho de columnas
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=clientes.xlsx'
        wb.save(response)
        return response

    paginator = Paginator(queryset, 20)  # 20 clientes por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Mostrar mensaje si no hay resultados y se aplicaron filtros
    mensaje_sin_resultados = None
    if not page_obj.object_list and filtros_aplicados:
        mensaje_sin_resultados = "No se encontraron clientes con los filtros aplicados."

    context = {
        'clientes': page_obj,
        'filtros': filtros,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'mensaje_sin_resultados': mensaje_sin_resultados,
    }

    return render(request, 'persona/lista_clientes.html', context)

@login_required
def cambiar_password_2(request):
    # Solo resetear la verificación si es la primera vez que se accede a la vista
    if request.method == 'GET' and not request.GET.get('verified'):
        request.session['password_verificada'] = False
    
    if request.method == 'POST':
        paso = request.POST.get('paso')
        
        if paso == 'verificar':
            password_actual = request.POST.get('password_actual')
            user = authenticate(username=request.user.username, password=password_actual)
            
            if user is not None:
                request.session['password_verificada'] = True
                messages.success(request, 'Contraseña verificada correctamente')
                return redirect(reverse('persona:cambiar_password_2') + '?verified=true')
            else:
                return redirect('persona:cambiar_password_2')
        
        elif paso == 'cambiar':
            if not request.session.get('password_verificada', False):
                messages.error(request, 'Debe verificar su contraseña actual primero')
                return redirect('persona:cambiar_password_2')
                
            password_nueva = request.POST.get('password_nueva')
            password_confirmar = request.POST.get('password_confirmar')
            
            if len(password_nueva) < 6 or len(password_nueva) > 16:
                messages.error(request, 'La contraseña debe estar entre 6 y 16 dígitos')
                return redirect(reverse('persona:cambiar_password_2') + '?verified=true')
            
            if password_nueva != password_confirmar:
                messages.error(request, 'Las contraseñas no coinciden')
                return redirect(reverse('persona:cambiar_password_2') + '?verified=true')
            
            request.user.set_password(password_nueva)
            request.user.save()
            update_session_auth_hash(request, request.user)
            messages.success(request, 'Contraseña cambiada exitosamente')
            # Después de cambiar la contraseña, resetear la verificación
            request.session['password_verificada'] = False
            
            # Redirigir según el tipo de usuario
            try:
                persona = Persona.objects.get(email=request.user.email)
                if persona.es_empleado or request.user.is_superuser:
                    return redirect('persona:inicio')
                else:
                    return redirect('persona:inicio')  # O cualquier otra vista para clientes
            except Persona.DoesNotExist:
                return redirect('persona:inicio')
    
    password_verificada = request.session.get('password_verificada', False)
    return render(request, 'persona/cambiar_password_2.html', {
        'password_verificada': password_verificada
    })

@empleado_requerido
@require_http_methods(["POST"])
def bloquear_cliente(request, persona_id):
    """
    Bloquea tanto el rol de cliente como el de empleado para un usuario.
    También cierra la sesión activa del usuario.
    """
    persona = get_object_or_404(Persona, id=persona_id)

    # Evitar que un empleado bloquee a un admin
    if persona.es_admin and not request.user.is_superuser:
        messages.error(request, 'No tienes permiso para bloquear a un administrador.')
        return JsonResponse({'status': 'error', 'message': 'No tienes permiso para bloquear a un administrador.'}, status=403)

    persona.bloqueado_cliente = True
    persona.bloqueado_empleado = True
    persona.save()
    
    # Forzar cierre de sesión del usuario afectado (la lógica se mantiene)
    try:
        user_afectado = User.objects.get(email=persona.email)
        from django.contrib.sessions.models import Session
        all_sessions = Session.objects.filter(expire_date__gte=timezone.now())
        for session in all_sessions:
            session_data = session.get_decoded()
            if str(session_data.get('_auth_user_id')) == str(user_afectado.id):
                session.delete()
    except (User.DoesNotExist, Exception):
        pass # No fallar si no se puede cerrar la sesión

    return JsonResponse({'status': 'success', 'message': f'El usuario {persona.email} ha sido bloqueado completamente.'})


@login_required
@require_http_methods(["POST"])
def desbloquear_cliente(request, persona_id):
    """
    Desbloquea SOLO el rol de cliente.
    Esta acción está restringida a administradores.
    """
    try:
        solicitante = Persona.objects.get(email=request.user.email)
        if not solicitante.es_admin:
            return JsonResponse({'status': 'error', 'message': 'No tienes los permisos de administrador para realizar esta acción.'}, status=403)
    except Persona.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'No se encontró tu perfil de usuario para verificar los permisos.'}, status=403)

    persona = get_object_or_404(Persona, id=persona_id)
    persona.bloqueado_cliente = False
    persona.save()
    
    return JsonResponse({'status': 'success', 'message': f'El rol de cliente para {persona.email} ha sido desbloqueado.'})

@empleado_requerido
def lista_empleados_gestion(request):
    """
    Vista para que los administradores gestionen a otros empleados (no-admins).
    """
    # Solo los admins pueden acceder
    if not request.user.is_superuser and (not hasattr(request.user, 'persona') or not request.user.persona.es_admin):
        messages.error(request, "No tienes permiso para acceder a esta página.")
        return redirect('persona:gestion')

    queryset = Persona.objects.filter(es_empleado=True, es_admin=False).order_by('apellido', 'nombre')

    filtros = {
        'nombre': request.GET.get('nombre', ''),
        'dni': request.GET.get('dni', ''),
        'email': request.GET.get('email', ''),
        'estado': request.GET.get('estado', ''),
        'fecha_desde': request.GET.get('fecha_desde', ''),
        'fecha_hasta': request.GET.get('fecha_hasta', ''),
    }

    if filtros['nombre']:
        queryset = queryset.filter(
            Q(nombre__icontains=filtros['nombre']) | Q(apellido__icontains=filtros['nombre'])
        )
    if filtros['dni']:
        queryset = queryset.filter(dni__icontains=filtros['dni'])
    if filtros['email']:
        queryset = queryset.filter(email__icontains=filtros['email'])
    if filtros['estado']:
        if filtros['estado'] == 'activo':
            queryset = queryset.filter(bloqueado_empleado=False)
        elif filtros['estado'] == 'bloqueado':
            queryset = queryset.filter(bloqueado_empleado=True)
    if filtros['fecha_desde']:
        queryset = queryset.filter(fecha_registro__gte=filtros['fecha_desde'])
    if filtros['fecha_hasta']:
        from datetime import datetime, timedelta
        fecha_hasta_dt = datetime.strptime(filtros['fecha_hasta'], '%Y-%m-%d') + timedelta(days=1)
        queryset = queryset.filter(fecha_registro__lt=fecha_hasta_dt)

    filtros_aplicados = any(f for f in filtros.values())

    if request.GET.get('export') == 'xlsx':
        import openpyxl
        from django.http import HttpResponse
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empleados"
        headers = ['Nombre', 'Apellido', 'Email', 'DNI', 'Fecha Nacimiento', 'Fecha Registro', 'Estado']
        ws.append(headers)
        for empleado in queryset:
            estado = "Bloqueado" if empleado.bloqueado_empleado else "Activo"
            fecha_nac_str = empleado.fecha_nacimiento.strftime('%d/%m/%Y') if empleado.fecha_nacimiento else 'N/A'
            ws.append([
                empleado.nombre, empleado.apellido, empleado.email,
                empleado.dni or 'N/A', fecha_nac_str,
                empleado.fecha_registro.strftime('%d/%m/%Y %H:%M'), estado
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=empleados.xlsx'
        wb.save(response)
        return response

    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    mensaje_sin_resultados = "No se encontraron empleados con los filtros aplicados." if not page_obj.object_list and filtros_aplicados else None

    context = {
        'empleados': page_obj,
        'filtros': filtros,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'mensaje_sin_resultados': mensaje_sin_resultados,
    }
    return render(request, 'persona/lista_empleados_gestion.html', context)


@require_http_methods(["POST"])
def bloquear_empleado(request, persona_id):
    if not request.user.is_superuser and (not hasattr(request.user, 'persona') or not request.user.persona.es_admin):
        return JsonResponse({'status': 'error', 'message': 'No tienes permiso para esta acción.'}, status=403)
    
    empleado = get_object_or_404(Persona, id=persona_id)
    empleado.bloqueado_empleado = True
    empleado.save()
    
    # Cerrar sesión del empleado afectado
    try:
        user_afectado = User.objects.get(email=empleado.email)
        from django.contrib.sessions.models import Session
        all_sessions = Session.objects.filter(expire_date__gte=timezone.now())
        for session in all_sessions:
            if str(session.get_decoded().get('_auth_user_id')) == str(user_afectado.id):
                session.delete()
    except:
        pass

    return JsonResponse({'status': 'success', 'message': f'El empleado {empleado.email} ha sido bloqueado.'})


@require_http_methods(["POST"])
def desbloquear_empleado(request, persona_id):
    if not request.user.is_superuser and (not hasattr(request.user, 'persona') or not request.user.persona.es_admin):
        return JsonResponse({'status': 'error', 'message': 'No tienes permiso para esta acción.'}, status=403)
        
    empleado = get_object_or_404(Persona, id=persona_id)
    empleado.bloqueado_empleado = False
    empleado.save()
    return JsonResponse({'status': 'success', 'message': f'El empleado {empleado.email} ha sido desbloqueado.'})

@empleado_requerido
@csrf_protect
@ensure_csrf_cookie
@require_http_methods(["GET", "POST"])
def registrar_cliente_nuevo(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            persona = form.save(commit=False)
            
            # Verificar que el email exista
            if not persona.email:
                form.add_error('email', 'El email es obligatorio para registrar un usuario.')
                return render(request, 'persona/registrar_cliente.html', {'form': form})
            
            # Generar contraseña aleatoria
            password = generar_password_random()
            
            try:
                # Crear usuario en Django
                user = User.objects.create_user(
                    username=persona.email,
                    email=persona.email,
                    password=password,
                    first_name=persona.nombre,
                    last_name=persona.apellido,
                    is_staff=False
                )
                
                persona.save()
                
                # Enviar email con la contraseña
                send_mail(
                    'Bienvenido a Alquil.Ar - Tu contraseña',
                    f'Hola {persona.nombre},\n\n'
                    f'Tu cuenta ha sido creada exitosamente.\n\n'
                    f'Credenciales de acceso:\n'
                    f'• Email: {persona.email}\n'
                    f'• Contraseña: {password}\n\n'
                    f'Por favor, cambia tu contraseña la próxima vez que inicies sesión.\n\n'
                    f'Tipo de usuario: Cliente\n\n'
                    f'Saludos,\n'
                    f'Equipo Alquil.Ar',
                    settings.DEFAULT_FROM_EMAIL,
                    [persona.email],
                    fail_silently=False,
                )
                
                messages.success(request, 'Cliente registrado exitosamente. Se ha enviado un email con las credenciales de acceso.')
                return redirect('persona:inicio')
            except Exception as e:
                # Si algo falla, eliminar el usuario si fue creado
                if 'user' in locals():
                    user.delete()
                form.add_error(None, f'Error al crear el usuario: {str(e)}')
                return render(request, 'persona/registrar_cliente.html', {'form': form})
    else:
        form = ClienteForm()
    return render(request, 'persona/registrar_cliente.html', {'form': form})

@empleado_requerido
@csrf_protect
@ensure_csrf_cookie
@require_http_methods(["GET", "POST"])
def registrar_empleado_nuevo(request):
    # Verificar si el usuario actual es admin
    if not request.user.is_superuser and (not hasattr(request.user, 'persona') or not request.user.persona.es_admin):
        messages.error(request, "No tienes permiso para registrar empleados.")
        return redirect('persona:gestion')

    if request.method == 'POST':
        form = EmpleadoForm(request.POST)
        if form.is_valid():
            persona = form.save(commit=False)
            
            # Verificar que el email exista
            if not persona.email:
                form.add_error('email', 'El email es obligatorio para registrar un empleado.')
                return render(request, 'persona/registrar_empleado_nuevo.html', {'form': form})
            
            # Generar contraseña aleatoria
            password = generar_password_random()
            
            try:
                # Crear usuario en Django
                user = User.objects.create_user(
                    username=persona.email,
                    email=persona.email,
                    password=password,
                    first_name=persona.nombre,
                    last_name=persona.apellido,
                    is_staff=True
                )
                
                persona.save()
                
                # Enviar email con la contraseña
                send_mail(
                    'Bienvenido a Alquil.Ar - Tu contraseña',
                    f'Hola {persona.nombre},\n\n'
                    f'Tu cuenta ha sido creada exitosamente como empleado.\n\n'
                    f'Credenciales de acceso:\n'
                    f'• Email: {persona.email}\n'
                    f'• Contraseña: {password}\n\n'
                    f'Por favor, cambia tu contraseña la próxima vez que inicies sesión.\n\n'
                    f'Tipo de usuario: Empleado\n\n'
                    f'Saludos,\n'
                    f'Equipo Alquil.Ar',
                    settings.DEFAULT_FROM_EMAIL,
                    [persona.email],
                    fail_silently=False,
                )
                
                messages.success(request, 'Empleado registrado exitosamente. Se ha enviado un email con las credenciales de acceso.')
                return redirect('persona:inicio')
            except Exception as e:
                # Si algo falla, eliminar el usuario si fue creado
                if 'user' in locals():
                    user.delete()
                form.add_error(None, f'Error al crear el usuario: {str(e)}')
                return render(request, 'persona/registrar_empleado_nuevo.html', {'form': form})
    else:
        form = EmpleadoForm()
    return render(request, 'persona/registrar_empleado_nuevo.html', {'form': form})

@require_GET
def estadisticas_maquinas(request):
    fecha_inicio = parse_date(request.GET.get('fecha_inicio'))
    fecha_fin = parse_date(request.GET.get('fecha_fin'))
    if not fecha_inicio or not fecha_fin or fecha_fin < fecha_inicio:
        return JsonResponse({'error': 'Fechas inválidas'}, status=400)
    # Filtrar alquileres por fecha de inicio en el rango
    alquileres = Alquiler.objects.filter(fecha_inicio__gte=fecha_inicio, fecha_inicio__lte=fecha_fin)
    # Agrupar por máquina base y contar
    ranking = (
        alquileres.values('maquina_base__nombre')
        .annotate(cantidad=Count('id'))
        .order_by('-cantidad')[:5]
    )
    labels = [item['maquina_base__nombre'] for item in ranking]
    cantidades = [item['cantidad'] for item in ranking]
    listado = [{'nombre': item['maquina_base__nombre'], 'cantidad': item['cantidad']} for item in ranking]
    return JsonResponse({'labels': labels, 'cantidades': cantidades, 'listado': listado})

# Create your views here.
